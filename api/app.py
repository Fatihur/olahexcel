import pandas as pd
from docx import Document
from flask import Flask, request, send_file, render_template
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Alignment, Border, Side, PatternFill, Font

app = Flask(__name__)

# Fungsi untuk mengolah data dari file Word
def process_word_data(docx_file):
    document = Document(docx_file)
    
    data = []
    current_name = ""
    total = 0
    row = [None] * 7  # Pastikan selalu ada 7 kolom
    
    # Fungsi untuk membersihkan dan mengonversi nominal menjadi float
    def convert_to_float(nominal_str):
        nominal_str = str(nominal_str)
        if isinstance(nominal_str, (int, float)):
            return float(nominal_str)
        
        cleaned_nominal = re.sub(r'[^0-9.]', '', nominal_str)  # Pastikan menjadi string
        try:
            return float(cleaned_nominal) if cleaned_nominal else 0.0
        except ValueError:
            return 0.0  # Jika gagal konversi, kembalikan 0.0

    # Fungsi untuk membersihkan dan mengonversi nominal menjadi string angka
    def clean_nominal(nominal_str):
        return re.sub(r'[^0-9]', '', nominal_str)

    # Loop untuk memproses setiap paragraf di dalam dokumen
    for para in document.paragraphs:
        text = para.text.strip()
        
        if "PERJALANAN" in text:
            continue
        
        if text and not any(keyword in text for keyword in ['Pesawat', 'By', 'Total']):  # Nama kru
            if current_name:  # Simpan data sebelumnya jika ada
                row[0] = len(data) + 1
                row[1] = current_name
                row[4] = "TOTAL"  # Tambahkan baris TOTAL di bawah nama
                row[5] = ""
                row[6] = total  # Masukkan total ke kolom TOTAL
                data.append(row)
                total = 0  # Reset total
            current_name = text  # Set nama baru
            row = [None] * 7  # Reset baris untuk data baru
        
        if 'By' in text:  # Data transaksi
            if current_name:
                nominal = text.split("=")[-1].strip()
                nominal_value = convert_to_float(nominal)
                row[0] = len(data) + 1
                row[1] = current_name
                row[4] = text
                row[5] = clean_nominal(nominal)  # Bersihkan nominal
                row[6] = ""
                data.append(row)
                total += nominal_value  # Menambahkan nilai nominal ke total
                row = [None] * 7  # Reset baris untuk transaksi berikutnya

    # Jika ada data terakhir yang belum disimpan
    if current_name:
        row[0] = len(data) + 1
        row[1] = current_name
        row[4] = "TOTAL"
        row[5] = ""
        row[6] = total  # Total untuk kru saat ini
        data.append(row)

    # Membuat DataFrame untuk disimpan dalam file Excel
    df = pd.DataFrame(data, columns=["NO", "NAMA", "POH", "TANGGAL NOTA", "DESKRIPSI", "NOMINAL", "TOTAL"])

    # Mengonversi kolom NOMINAL menjadi angka sebelum menyimpan ke Excel
    df['NOMINAL'] = df['NOMINAL'].apply(lambda x: convert_to_float(x))

    # Menyimpan data ke dalam file Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Processed Data')
    
    # Mengakses workbook untuk menambahkan rumus
    output.seek(0)
    wb = load_workbook(output)
    ws = wb['Processed Data']

    # Menambahkan rumus SUM di kolom TOTAL untuk setiap grup NAMA
    row_start = 2  # Baris mulai setelah header
    current_name = None
    start_row = row_start
    for row in range(row_start, len(df) + 2):  # Menambah 2 karena header dan Excel dimulai dari baris 1
        if ws[f'B{row}'].value != current_name:
            if current_name is not None:
                # Menambahkan rumus SUM untuk grup sebelumnya
                ws[f'G{row-1}'] = f'=SUM(F{start_row}:F{row-1})'
            current_name = ws[f'B{row}'].value
            start_row = row
    
    # Menambahkan rumus SUM untuk grup terakhir
    ws[f'G{row-1}'] = f'=SUM(F{start_row}:F{row-1})'

    # Merge the 'TOTAL' cell with the adjacent cell in the Excel file
    for row in range(row_start, len(df) + 2):
        if ws[f'E{row}'].value == "TOTAL":
            ws[f'E{row}'] = "TOTAL ="
            ws.merge_cells(f'E{row}:F{row}')
            ws[f'E{row}'].alignment = Alignment(horizontal='center')

    # Mengatur format angka di kolom NOMINAL (kolom F) tanpa angka di belakang koma
    # Membuat style untuk angka tanpa desimal
    number_style = NamedStyle(name="number_style", number_format="#,##0")
    
    # Terapkan format angka ke setiap sel di kolom F (NOMINAL)
    for row in range(2, len(df) + 2):  # Dimulai dari baris ke-2 setelah header
        ws[f'F{row}'].style = number_style

    # Merge cells in column 'B' (NAMA) and repeat 'NO' for each name
    current_name = None
    start_merge_row = None
    current_no = 1
    for row in range(2, len(df) + 2):
        name_cell = ws[f'B{row}']
        no_cell = ws[f'A{row}']
        if name_cell.value == current_name:
            # If name is the same as the previous one, merge the cells
            ws.merge_cells(f'B{start_merge_row}:B{row}')
            ws.merge_cells(f'A{start_merge_row}:A{row}')
            name_cell.alignment = Alignment(horizontal='center')
            no_cell.value = current_no
        else:
            current_name = name_cell.value
            start_merge_row = row
            no_cell.value = current_no
            current_no += 1

    # Set middle alignment for 'NAMA' and 'NO' columns
    middle_alignment = Alignment(vertical='center', horizontal='center')
    for row in range(2, len(df) + 2):
        ws[f'A{row}'].alignment = middle_alignment
        ws[f'B{row}'].alignment = middle_alignment

    # Enhance the Excel sheet by adding borders, background colors, and adjusting column widths for better presentation
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_font = Font(bold=True)

    # Adjust column widths
    for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[column].width = 15

    # Apply styles to header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border

    # Apply styles to all cells
    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border

    # Add a grand total row at the bottom of the 'TOTAL' column
    grand_total_row = len(df) + 3  # Account for the header and 1-based index
    ws[f'F{grand_total_row}'] = 'Grand Total'
    ws[f'G{grand_total_row}'] = f'=SUM(G2:G{grand_total_row-1})'
    ws[f'F{grand_total_row}'].alignment = middle_alignment
    ws[f'G{grand_total_row}'].alignment = middle_alignment
    ws[f'F{grand_total_row}'].font = header_font
    ws[f'G{grand_total_row}'].font = header_font

    # Menyimpan file Excel yang telah dimodifikasi dengan rumus dan format
    new_output = io.BytesIO()
    wb.save(new_output)
    new_output.seek(0)

    return new_output

# Fungsi untuk mengunduh file Excel
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            output = process_word_data(file)
            # Extract the original file name without extension and append '_processed.xlsx'
            original_filename = file.filename.rsplit('.', 1)[0]
            download_filename = f"{original_filename}_processed.xlsx"

            # Mengirim file Excel untuk diunduh dengan nama yang sesuai
            return send_file(output, as_attachment=True, download_name=download_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
